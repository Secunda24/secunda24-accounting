import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.core.config import CONFIG_PATH, settings
from app.models.schemas import ReceiptRow, StatementRow


LEDGER_HEADERS = [
    "ID",
    "Date",
    "Description",
    "Category",
    "Debit",
    "Credit",
    "Balance",
    "Source Type",
    "Source File",
    "Notes",
    "Status",
]

RECEIPT_HEADERS = [
    "Receipt ID",
    "Date",
    "Vendor",
    "Amount",
    "Tax",
    "Linked Ledger ID",
    "Source File",
    "OCR Confidence",
    "Notes",
]

IMPORT_HEADERS = [
    "Import Date/Time",
    "File Name",
    "File Type",
    "Rows Extracted",
    "Result",
    "Error Notes",
]

SETTINGS_HEADERS = ["Key", "Value"]


def ensure_workbook() -> Path:
    workbook_path = Path(settings.workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    Path(settings.exports_dir).mkdir(parents=True, exist_ok=True)

    if workbook_path.exists():
        return workbook_path

    workbook = Workbook()
    workbook.remove(workbook.active)

    ledger = workbook.create_sheet("Ledger")
    receipts = workbook.create_sheet("Receipts")
    imports_log = workbook.create_sheet("Imports Log")
    settings_sheet = workbook.create_sheet("Settings")

    ledger.append(LEDGER_HEADERS)
    receipts.append(RECEIPT_HEADERS)
    imports_log.append(IMPORT_HEADERS)
    settings_sheet.append(SETTINGS_HEADERS)
    settings_sheet.append(["config_file", str(CONFIG_PATH)])
    settings_sheet.append(["currency", settings.currency])
    settings_sheet.append(["categories", ", ".join(settings.default_categories)])
    settings_sheet.append(["bank_parser", settings.bank_parser])
    settings_sheet.append(["tesseract_cmd", settings.tesseract_cmd])
    settings_sheet.append(["poppler_path", settings.poppler_path])

    workbook.save(workbook_path)
    return workbook_path


def load_local_workbook():
    return load_workbook(ensure_workbook())


def _next_id(sheet, id_column_index: int) -> int:
    values = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        value = row[id_column_index - 1]
        if value:
            try:
                values.append(int(str(value).split("-")[-1]))
            except ValueError:
                continue
    return max(values, default=0) + 1


def append_statement_rows(rows: list[StatementRow]) -> int:
    workbook = load_local_workbook()
    sheet = workbook["Ledger"]
    next_id = _next_id(sheet, 1)
    for offset, row in enumerate(rows):
        sheet.append(
            [
                f"L-{next_id + offset:05d}",
                row.date,
                row.description,
                row.category,
                row.debit,
                row.credit,
                row.balance,
                "Statement",
                row.source_file,
                row.notes,
                row.status,
            ]
        )
    workbook.save(settings.workbook_path)
    return len(rows)


def append_receipt_rows(rows: list[ReceiptRow]) -> int:
    workbook = load_local_workbook()
    sheet = workbook["Receipts"]
    next_id = _next_id(sheet, 1)
    for offset, row in enumerate(rows):
        receipt_id = row.receipt_id or f"R-{next_id + offset:05d}"
        sheet.append(
            [
                receipt_id,
                row.date,
                row.vendor,
                row.amount,
                row.tax,
                row.linked_ledger_id,
                row.source_file,
                row.ocr_confidence,
                row.notes,
            ]
        )
    workbook.save(settings.workbook_path)
    return len(rows)


def log_import(file_name: str, file_type: str, rows_extracted: int, result: str, error_notes: str = "") -> None:
    workbook = load_local_workbook()
    sheet = workbook["Imports Log"]
    sheet.append(
        [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            file_name,
            file_type,
            rows_extracted,
            result,
            error_notes,
        ]
    )
    workbook.save(settings.workbook_path)


def get_sheet_rows(sheet_name: str) -> list[dict]:
    workbook = load_local_workbook()
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def filter_ledger(filters: dict) -> list[dict]:
    rows = get_sheet_rows("Ledger")
    filtered = []
    for row in rows:
        row_date = str(row.get("Date") or "")
        if filters.get("date_from") and row_date and row_date < filters["date_from"]:
            continue
        if filters.get("date_to") and row_date and row_date > filters["date_to"]:
            continue
        if filters.get("category") and row.get("Category") != filters["category"]:
            continue
        if filters.get("source") and row.get("Source Type") != filters["source"]:
            continue
        amount = max(float(row.get("Debit") or 0), float(row.get("Credit") or 0))
        if filters.get("min_amount") not in (None, "") and amount < float(filters["min_amount"]):
            continue
        if filters.get("max_amount") not in (None, "") and amount > float(filters["max_amount"]):
            continue
        filtered.append(row)
    return filtered


def export_workbook_copy() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = Path(settings.exports_dir) / f"ledger_export_{timestamp}.xlsx"
    shutil.copy2(settings.workbook_path, target)
    return target


def sync_settings_sheet() -> None:
    workbook = load_local_workbook()
    if "Settings" in workbook.sheetnames:
        workbook.remove(workbook["Settings"])

    sheet = workbook.create_sheet("Settings")
    sheet.append(SETTINGS_HEADERS)
    sheet.append(["config_file", str(CONFIG_PATH)])
    sheet.append(["currency", settings.currency])
    sheet.append(["categories", ", ".join(settings.default_categories)])
    sheet.append(["bank_parser", settings.bank_parser])
    sheet.append(["tesseract_cmd", settings.tesseract_cmd])
    sheet.append(["poppler_path", settings.poppler_path])
    workbook.save(settings.workbook_path)


def get_recent_imports(limit: int = 5) -> list[dict]:
    rows = get_sheet_rows("Imports Log")
    return list(reversed(rows))[:limit]
